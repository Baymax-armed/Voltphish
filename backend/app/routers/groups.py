"""Recipient group CRUD. All routes require authentication (A01)."""
from __future__ import annotations

from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session as DbSession

from ..database import get_db
from ..csrf import csrf_protect
from ..dependencies import get_current_user
from ..models import Group, Target
from ..schemas.common import Message
from ..schemas.group import GroupCreate, GroupOut, GroupSummary, GroupUpdate

router = APIRouter(
    prefix="/api/v1/groups",
    tags=["groups"],
    dependencies=[Depends(get_current_user), Depends(csrf_protect)],
)


@router.get("", response_model=list[GroupSummary])
def list_groups(db: DbSession = Depends(get_db)) -> list[GroupSummary]:
    groups = db.execute(select(Group).order_by(Group.name)).scalars()
    return [
        GroupSummary(
            id=g.id, name=g.name, target_count=len(g.targets), modified_at=g.modified_at
        )
        for g in groups
    ]


@router.post("", response_model=GroupOut, status_code=status.HTTP_201_CREATED)
def create_group(payload: GroupCreate, db: DbSession = Depends(get_db)) -> Group:
    group = Group(name=payload.name)
    group.targets = [
        Target(
            email=str(t.email).lower(),
            phone=t.phone,
            first_name=t.first_name,
            last_name=t.last_name,
            position=t.position,
            is_vip=t.is_vip,
        )
        for t in payload.targets
    ]
    db.add(group)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status.HTTP_409_CONFLICT, "A group with that name already exists")
    db.refresh(group)
    return group


def _rows_to_targets(rows: list[list[object]]) -> list[dict]:
    """Turn spreadsheet rows into target dicts, mirroring the CSV importer:
    honor a header row (Email/First/Last/Position/Phone/VIP, any order) else
    auto-detect the email column by its '@'. Dedupes by lower-cased email."""
    grid = [[("" if c is None else str(c)).strip() for c in r] for r in rows]
    grid = [r for r in grid if any(r)]
    if not grid:
        return []

    first = [c.lower() for c in grid[0]]
    has_header = any("email" in c for c in first) and not any("@" in c for c in first)
    idx = {"email": -1, "first": -1, "last": -1, "position": -1, "phone": -1, "vip": -1}
    data = grid
    if has_header:
        for i, c in enumerate(first):
            if "email" in c:
                idx["email"] = i
            elif "first" in c:
                idx["first"] = i
            elif "last" in c:
                idx["last"] = i
            elif any(k in c for k in ("position", "title", "department", "role")):
                idx["position"] = i
            elif "phone" in c or "mobile" in c:
                idx["phone"] = i
            elif "vip" in c:
                idx["vip"] = i
        data = grid[1:]

    def cell(cols: list[str], i: int) -> str | None:
        return cols[i] if 0 <= i < len(cols) and cols[i] else None

    truthy = {"1", "y", "yes", "true", "vip"}
    out: list[dict] = []
    seen: set[str] = set()
    for cols in data:
        if has_header and idx["email"] >= 0:
            email = cell(cols, idx["email"])
            first_name = cell(cols, idx["first"])
            last_name = cell(cols, idx["last"])
            position = cell(cols, idx["position"])
            phone = cell(cols, idx["phone"])
            vip_raw = cell(cols, idx["vip"])
            is_vip = bool(vip_raw) and vip_raw.lower() in truthy
        else:
            email = next((c for c in cols if "@" in c), None)
            if not email:
                continue
            rest = [c for c in cols if c != email]
            first_name = rest[0] if len(rest) > 0 and rest[0] else None
            last_name = rest[1] if len(rest) > 1 and rest[1] else None
            position = rest[2] if len(rest) > 2 and rest[2] else None
            phone = None
            is_vip = False
        if not email or "@" not in email:
            continue
        el = email.lower()
        if el in seen:
            continue
        seen.add(el)
        out.append({
            "email": el, "first_name": first_name, "last_name": last_name,
            "position": position, "phone": phone, "is_vip": is_vip,
        })
    return out


@router.post("/parse-xlsx")
async def parse_xlsx(file: UploadFile = File(...)) -> list[dict]:
    """Parse an uploaded .xlsx target list into rows the client merges into the
    group editor. Done server-side because xlsx is a binary format; openpyxl
    reads it in read-only, data-only mode (no formulas/macros run)."""
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, "File exceeds 5 MB.")
    try:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)] if ws is not None else []
        wb.close()
    except Exception:  # noqa: BLE001 — corrupt / not an xlsx
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Couldn't read that .xlsx file.")
    return _rows_to_targets(rows)


@router.get("/{group_id}", response_model=GroupOut)
def get_group(group_id: int, db: DbSession = Depends(get_db)) -> Group:
    group = db.get(Group, group_id)
    if group is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Group not found")
    return group


@router.put("/{group_id}", response_model=GroupOut)
def update_group(group_id: int, payload: GroupUpdate, db: DbSession = Depends(get_db)) -> Group:
    group = db.get(Group, group_id)
    if group is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Group not found")
    group.name = payload.name
    # Replace membership wholesale (cascade delete-orphan cleans old rows).
    group.targets = [
        Target(
            email=str(t.email).lower(),
            phone=t.phone,
            first_name=t.first_name,
            last_name=t.last_name,
            position=t.position,
            is_vip=t.is_vip,
        )
        for t in payload.targets
    ]
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status.HTTP_409_CONFLICT, "A group with that name already exists")
    db.refresh(group)
    return group


@router.delete("/{group_id}", response_model=Message)
def delete_group(group_id: int, db: DbSession = Depends(get_db)) -> Message:
    group = db.get(Group, group_id)
    if group is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Group not found")
    db.delete(group)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status.HTTP_409_CONFLICT, "This group is still used by a campaign — delete those campaigns first.")
    return Message(detail="Group deleted")
